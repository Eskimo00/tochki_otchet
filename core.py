from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import timedelta
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side

# Регулярные выражения для поиска данных
MODEL_RE = re.compile(r"Модель:\s*(.+?)(?:\s+Номер ТС:.*)?$", re.IGNORECASE)
PERIOD_RE = re.compile(r"Период:\s*с\s*(\d{2}\.\d{2}\.\d{4})", re.IGNORECASE)
POINT_RE = re.compile(r"Точка\s*(\d+)", re.IGNORECASE)

# Стили оформления
FILL_HEADER = PatternFill("solid", fgColor="FFF2CC")  # желтый для заголовков
FILL_BODY = PatternFill("solid", fgColor="FFFBEA")  # светло-желтый фон таблицы
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")  # зелёный для точек с проходами
FILL_RED = PatternFill("solid", fgColor="FFC7CE")  # красный для "нет"
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


@dataclass
class ModelResult:
    """Хранит агрегированные данные по модели."""

    name: str
    group: str
    total_points: Optional[int]
    counts: dict[int, int]
    sums: dict[int, timedelta]


def parse_duration(value: object) -> timedelta:
    """Парсим строку вида 00:05:12 в timedelta."""
    if value is None:
        return timedelta(0)
    s = str(value).strip()
    parts = s.split(":")
    if len(parts) != 3:
        return timedelta(0)
    try:
        h, m, sec = (int(float(p)) for p in parts)
    except ValueError:
        return timedelta(0)
    return timedelta(hours=h, minutes=m, seconds=sec)


def format_duration(td: timedelta) -> str:
    total_seconds = int(td.total_seconds())
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def detect_columns(ws) -> Tuple[int, int]:
    """Определяем номера колонок 'Название КТ' и 'Продолжительность' по заголовкам."""
    name_col = 3
    dur_col = 10
    max_scan_rows = min(40, ws.max_row)
    max_scan_cols = min(30, ws.max_column)
    for r in range(1, max_scan_rows + 1):
        for c in range(1, max_scan_cols + 1):
            text = ws.cell(r, c).value
            if not text:
                continue
            txt = str(text).strip().lower()
            if "название" in txt:
                name_col = c
            if "продолж" in txt:
                dur_col = c
    return name_col, dur_col


def ensure_unique_path(path: Path) -> Path:
    """Если файл существует — добавляем (1), (2) и т.д."""
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    i = 1
    while True:
        candidate = parent / f"{stem} ({i}){suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def build_default_output(source_path: Path) -> Path:
    """Строим путь по умолчанию рядом с исходным файлом."""
    base_dir = source_path.parent if source_path else Path.cwd()
    return ensure_unique_path(base_dir / "отчет прохождения точек.xlsx")


def _extract_period(texts: Iterable[str]) -> Optional[str]:
    for txt in texts:
        match = PERIOD_RE.search(txt)
        if match:
            return match.group(1)
    return None


def _extract_model_name(texts: Iterable[str]) -> Optional[str]:
    for txt in texts:
        match = MODEL_RE.search(txt)
        if match:
            return match.group(1).strip()
    return None


def _parse_model_info(model_name: str) -> tuple[str, str, Optional[int]]:
    """Возвращает (группа, название без счетчика и префикса, всего точек)."""
    name = model_name.strip()
    total_points: Optional[int] = None

    points_match = re.search(r"\((\d+)\)\s*$", name)
    if points_match:
        total_points = int(points_match.group(1))
        name = name[: points_match.start()].rstrip()

    group_match = re.match(r"(ТЭ|ТАН-\d+)", name, re.IGNORECASE)
    group = group_match.group(1).upper() if group_match else (name.split()[0] if name else "")

    if group:
        name = re.sub(rf"^\s*{re.escape(group)}\s*[-:]?\s*", "", name, flags=re.IGNORECASE)

    return group, name, total_points


def _extract_point_index(text: str) -> Optional[int]:
    match = POINT_RE.search(text)
    if not match:
        return None
    idx = int(match.group(1))
    if 1 <= idx <= 8:
        return idx
    return None


def _is_section_end(texts: Iterable[str]) -> bool:
    return any("ИТОГО по ТС" in txt for txt in texts)


def _create_empty_result(model_name: str) -> ModelResult:
    group, clean_name, total_points = _parse_model_info(model_name)
    return ModelResult(
        name=clean_name,
        group=group,
        total_points=total_points,
        counts={i: 0 for i in range(1, 9)},
        sums={i: timedelta(0) for i in range(1, 9)},
    )


def _update_point_data(
    result: ModelResult,
    point_text: str,
    duration_text: str,
) -> None:
    point_index = _extract_point_index(point_text)
    if point_index is None:
        return
    result.counts[point_index] += 1
    result.sums[point_index] += parse_duration(duration_text)


def _auto_width(ws) -> None:
    """Автоподбор ширины колонок."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            val = cell.value
            if val is None:
                continue
            length = len(str(val))
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def _build_workbook(results: List[ModelResult], period_date: str) -> Workbook:
    """Создаём Excel-файл с итогами."""
    out_wb = Workbook()
    out_ws = out_wb.active
    headers = ["№ п/п", "Хозяйство", "Название объекта", "Всего точек", "Дата"] + [
        f"Точка {i}" for i in range(1, 9)
    ]
    out_ws.append(headers)

    header_font = Font(bold=True)
    for cell in out_ws[1]:
        cell.fill = FILL_HEADER
        cell.border = THIN_BORDER
        cell.font = header_font

    for row_idx, result in enumerate(results, start=2):
        line_no = row_idx - 1
        line = [line_no, result.group, result.name, result.total_points or "", period_date]
        point_values = []
        for point_index in range(1, 9):
            count = result.counts[point_index]
            total_duration = result.sums[point_index]
            if count == 0:
                point_values.append("---")
            elif total_duration.total_seconds() == 0:
                point_values.append("нет")
            else:
                duration = format_duration(total_duration)
                point_values.append(f"({count}) / {duration}")
        line.extend(point_values)
        out_ws.append(line)

        for col_idx, cell in enumerate(out_ws[row_idx], start=1):
            cell.border = THIN_BORDER
            if col_idx >= 6 or (col_idx < 6 and row_idx > 1):
                cell.fill = FILL_BODY
            if col_idx >= 6:
                value = cell.value
                cell.fill = FILL_RED if value in ("нет", "---") else FILL_GREEN

    # Объединяем ячейки по хозяйству для наглядности
    group_col = 2
    start_row = 2
    current_group = out_ws.cell(row=start_row, column=group_col).value if out_ws.max_row >= 2 else None
    for r in range(3, out_ws.max_row + 1):
        value = out_ws.cell(row=r, column=group_col).value
        if value != current_group:
            if current_group and r - start_row > 1:
                out_ws.merge_cells(start_row=start_row, end_row=r - 1, start_column=group_col, end_column=group_col)
            start_row = r
            current_group = value
    if current_group and out_ws.max_row - start_row + 1 > 1:
        out_ws.merge_cells(start_row=start_row, end_row=out_ws.max_row, start_column=group_col, end_column=group_col)

    _auto_width(out_ws)
    return out_wb


def process_file(source: Path, dest: Path) -> Path:
    """Читает исходный Excel, агрегирует проходы и сохраняет новый файл."""
    wb = load_workbook(source, data_only=True)
    ws = wb.active

    name_col, dur_col = detect_columns(ws)

    period_date = ""
    results: List[ModelResult] = []
    current_model: Optional[ModelResult] = None

    def flush_model() -> None:
        nonlocal current_model
        if current_model is None:
            return
        results.append(current_model)
        current_model = None

    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        texts = [str(v).strip() if v is not None else "" for v in row_vals]

        period = _extract_period(texts)
        if period:
            period_date = period

        model_name = _extract_model_name(texts)
        if model_name:
            flush_model()
            current_model = _create_empty_result(model_name)

        if current_model:
            name_txt = texts[name_col - 1] if name_col - 1 < len(texts) else ""
            dur_txt = texts[dur_col - 1] if dur_col - 1 < len(texts) else ""
            _update_point_data(current_model, name_txt, dur_txt)

        if _is_section_end(texts):
            flush_model()

    flush_model()

    group_order = {"ТЭ": 0, "ТАН-1": 1, "ТАН-2": 2, "ТАН-3": 3}

    results.sort(
        key=lambda item: (
            group_order.get(item.group.upper(), 100),
            item.group.lower(),
            item.name.lower(),
        )
    )
    out_wb = _build_workbook(results, period_date)

    dest = ensure_unique_path(dest)
    out_wb.save(dest)
    return dest
